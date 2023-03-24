import tkinter as tk
import openpyxl
import re


class TABigDictionary(object):
    def __init__(self):
        self.root = tk.Tk()
        self.root.iconbitmap('./icon/TA_大字典.ico')
        self.root.title("TA 大字典 o.1版")
        self.search_value = tk.StringVar()

    def show_information(self, event, worksheet, output_notebook, output_notebook_2):
        """

        :param event: 鼠标点击事件
        :param worksheet: 大字典的源表格加载得到的 Sheet
        :param output_notebook: 根据搜索框的输入过滤得到的数值，默认全部显示
        :param output_notebook_2: 根据鼠标在搜索框内点击，显示指定输出结果的位置
        :return: None
        """
        # 在函数调用显示之前，先清除之前的显示记录
        output_notebook_2.configure(state='normal')
        output_notebook_2.delete('1.0', tk.END)
        output_notebook_2.configure(state="disabled")

        # 获取点击行的索引
        index = int(event.widget.index(tk.CURRENT).split('.')[0]) - 1

        # 获取该行所有文本信息
        line_start = f"{index + 1}.0"
        line_end = f"{index + 1}.end"
        data = output_notebook.get(line_start, line_end)

        # 对output_notebook获取的文本信息进行提取 “名称” 这个一关键语句
        pattern = r'\S+\s+(\S+).*'
        match = re.match(pattern, data)
        data_re_result = ""
        if match:
            data_re_result = match.group(1)

        # 在源表格内遍历数据
        for hang in range(1, worksheet.max_row+1):
            cell_name = "B" + str(hang)
            if data_re_result == worksheet[cell_name].value:
                data = worksheet["D" + str(hang)].value

        # 遇见句号换行，且保留句号
        data_line_break = re.sub(r'\。', '。\n', data)

        # 将data的数据显示在第二个文字框内
        output_notebook_2.configure(state='normal', font=("Black", 13), spacing1=15)
        output_notebook_2.insert(tk.END, str(data_line_break) + '\n')
        output_notebook_2.configure(state='disabled')

    def workbook_show(self, entry_value, output_notebook, output_notebook_2):
        """

        :param entry_value: UI内输入框输入要搜索的内容
        :param output_notebook: 根据搜索框的输入过滤得到的数值，默认全部显示
        :param output_notebook_2: 根据鼠标在搜索框内点击，显示指定输出结果的位置
        :return: None
        """
        output_notebook.configure(state='normal')
        output_notebook.delete('1.0', tk.END)
        output_notebook.configure(state="disabled")

        # 表格处理
        workbook = openpyxl.load_workbook("./xlsx文件夹/TA大字典.xlsx")

        # 选择一个工作表
        worksheet = workbook.active

        # 将工作表中的数据格式化为字符串
        for i, row in enumerate(worksheet.iter_rows(values_only=True)):
            if not entry_value or entry_value in row:  # 搜索值为空的时候能全部显示，当有搜索值的时候只显示搜索值
                row_str = ''
                for cell in row[:3]:
                    cell_str = str(cell).ljust(25)  # 将数据左对齐，并占用25个字符的宽度
                    row_str += cell_str
                # 在文本框中插入表格数据
                output_notebook.configure(state='normal')
                output_notebook.insert(tk.END, row_str + '\n', f"row{i}")
                output_notebook.tag_bind(f"row{i}", "<Button-1>", lambda event: self.show_information(event, worksheet, output_notebook, output_notebook_2))
                output_notebook.configure(state='disabled')
        # 设置选中后的背景色为红色
        output_notebook.tag_configure("sel", background="red")

    def creating_ui(self):
        self.root.geometry("600x700+100+100")

        # 开始对整个面板进行分区处理
        frame = tk.PanedWindow(self.root, orient='vertical', sashrelief='sunken')
        frame.pack(fill='both', expand=True)
        frame_up = tk.PanedWindow(self.root, orient='horizontal', sashrelief='sunken')
        frame_dw = tk.PanedWindow(self.root, orient='horizontal', sashrelief='sunken')
        frame_text = tk.PanedWindow(self.root, orient='horizontal', sashrelief='sunken')
        frame.add(frame_up, height=35), frame.add(frame_dw, height=265), frame.add(frame_text, height=300)

        # 第一面板的参数
        search_label = tk.Label(frame_up, text='请准确输入：')
        search_entry = tk.Entry(frame_up, textvariable=self.search_value)
        search_enter = tk.Button(frame_up, text='搜索', font=('楷体', 12), fg='black', width=5, height=1, command=lambda: self.workbook_show(search_entry.get(), output_notebook, output_notebook_2))
        # 第一面板参数排列
        search_label.grid(row=1, column=1)
        search_entry.grid(row=1, column=2)
        search_enter.grid(row=1, column=3)

        # 在下方Text窗口输出表格信息，并支持点击某一行输出信息。
        output_notebook = tk.Text(frame_dw, state="disabled")
        output_scrollbar = tk.Scrollbar(frame_dw)
        output_scrollbar.config(command=output_notebook.yview)  # 下拉式的菜单栏
        output_notebook.config(yscrollcommand=output_scrollbar.set)

        output_notebook.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        output_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 最下面的文字输出展示框
        output_notebook_2 = tk.Text(frame_text, state="disabled")
        output_scrollbar_2 = tk.Scrollbar(frame_text)
        output_scrollbar_2.config(command=output_notebook_2.yview)
        output_notebook_2.config(yscrollcommand=output_scrollbar_2.set)

        output_notebook_2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        output_scrollbar_2.pack(side=tk.RIGHT, fill=tk.Y)

        self.workbook_show(search_entry.get(), output_notebook, output_notebook_2)

        self.root.mainloop()


if __name__ == '__main__':
    TABigDictionary().creating_ui()
